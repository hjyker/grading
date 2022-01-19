import glob
import os
from pathlib import Path

from openpyxl import load_workbook


def parse_origin_data(data_path):
    for file in glob.glob(f'{data_path}/*_reviews.xlsx'):
        review_path = Path(file)
        music_no = review_path.name.split('_')[0]
        label_path = review_path.parent / f'{music_no}_labels.xlsx'
        if label_path.exists():
            review_file = load_workbook(review_path)['Sheet1']
            label_file = load_workbook(label_path)['Sheet1']
            reviews = [(aa.value, bb.value)
                       for aa, bb in review_file.rows if aa.value is not None][1:]
            labels = [lb[0].value for lb in label_file.rows][1:]
            yield music_no, reviews, labels
        else:
            print(f'ERROR no file {label_path}')


def create_comment(music_no, reviews):
    comments = [Comments(music_no=music_no, comment=review, likes=int(likes))
                for review, likes in reviews]
    try:
        Comments.objects.bulk_create(comments)
    except Exception:
        print(comments)


def create_label(music_no, labels):
    labelss = [Labels(music_no=music_no, label=label) for label in labels]
    Labels.objects.bulk_create(labelss)


def create_order(music_no):
    AuditOrder.objects.create(music_no=music_no)


if __name__ == '__main__':
    import django
    os.sys.path.extend(['/code'])
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'findiff.settings')
    django.setup()

    from findiff.models import AuditOrder, Comments, Labels

    for music_no, reviews, labels in parse_origin_data('./init_dataset'):
        create_comment(music_no, reviews)
        create_label(music_no, labels)
        create_order(music_no)
